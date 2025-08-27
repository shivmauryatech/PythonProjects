# This page contains the main function to run the banking system.

from bank import Account
from openpyxl import load_workbook

def superuser_view():
    wb = load_workbook("accounts.xlsx")
    sheet = wb["Ledger"]
    print("\nSuperuser View of All Transactions:")
    for row in sheet.iter_rows(values_only=True):
        print(row)
    

def main():
    print("\n\n***************************************************************")
    print("**************Welcome to Shiva Banking System******************")
    print("***************************************************************\n")
    while True:
        print('\n Choose and option from below menu:\n ')
        print("1. Open New Account ")
        print("2. Credit Amount ")
        print("3. Debit Amount ")
        print("4. Check Balance ")
        print("5. Superuser View (All Transactions) ")
        print("6. Exit ")

        choice = input("++++++++++Enter your choice (1-6): +++++++++++++\n")

        if choice == '1':
            name = input("Enter Account Holder Name: ")
            acc = Account(name)
            print(f"Account created for {name}.")
        elif choice in ['2', '3', '4']:
            name = input("Enter Account Holder Name: ")
            acc = Account(name)
            if choice == '2':
                amount = input("Enter amount to credit: ")
                acc.credit(amount)
            elif choice == '3':
                amount = input("Enter amount to debit: ")
                acc.debit(amount)
            elif choice == '4':
                balance = acc.get_balance()
                print(f"Current balance for {name} is: {balance}")
        elif choice == '5':
            password = input("Enter superuser password :  ")
            if password == 'admin123':
                superuser_view()
            else:
                print("Invalid Password")
        
        elif choice == '6':
            print(" *****************************************")
            print(" ** Thanks for banking with Shiva Bank! **")
            print(" *****************************************")
            break
        else:
            print(" Invalid choice, Please try again ")
    
if __name__ =='__main__':
    main()
