""" This is the abstract class for all types of bank accounts. This implements the basic
functionalities of a bank account. """

from abc import ABC, abstractmethod
from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = "accounts.xlsx"

class BankAccount(ABC):
    @abstractmethod
    def credit(self,amount):
        pass

    @abstractmethod
    def debit(self,amount):
        pass
    @abstractmethod
    def get_balance(self):
        pass

class Account(BankAccount):
    def __init__(self,name):
        self.name = name
        self.sheet = self.load_or_create_sheet()
    
    def load_or_create_sheet(self):
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Ledger"
            ws.append(["Name               |", "Type   |", "Amount"])
            wb.save(EXCEL_FILE)
        return load_workbook(EXCEL_FILE)["Ledger"]
    
    def _save_transaction(self, type_, amount):
        self.sheet.append([self.name, type_, amount])
        self.sheet.parent.save(EXCEL_FILE)

    def credit(self, amount):
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("\nAmount must be positive.")
            self._save_transaction("Credit", amount)
            print(f"*******Credited {amount} to {self.name}'s account.*******")
        except Exception as e:
            print(f"Error: {e}")

    def debit(self, amount):
        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("\nAmount must be positive.")
            if amount > self.get_balance():
                raise ValueError("\n Insufficient funds.")
            self._save_transaction("Debit", amount)
            print(f"*******Debited {amount} from {self.name}'s account.**********")
        except Exception as e:
            print(f"Error: {e}")

    def get_balance(self):
        balance = 0
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == self.name:
                balance += row[2] if row[1] == "Credit" else -row[2]
        return balance